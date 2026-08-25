import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURACION GENERAL Y UMBRALES
# ============================================================
IP_SV = "192.168.0.178"
MAX_SEGMENTOS = 20000
UMBRAL_INSTANTANEO_S = 0.001  # 1 milisegundo

# ============================================================
# DEFINICION DE LA TOPOLOGIA Y RUTAS (Solo PC1 y PC3)
# ============================================================
# IMPORTANTE: Reemplaza YYY por la IP real del PC3
TOPOLOGIA = [
    {
        "equipo": "PC1",
        "ip": "192.168.0.148",
        "archivo_salida": "TEST_1_PC1_Resultados.xlsx",
        "pruebas": [
            {"id": "P1", "nombre": "Prueba 1", "pc_csv": r"D:\Archivos de Programa\CapturasWS\PC1_TELETRAFICO\TEST 1\PC1T1_1.csv", "sv_csv": r"D:\Archivos de Programa\CapturasWS\SV\TEST 1\SVT1_P1.csv"},
            {"id": "P2", "nombre": "Prueba 2", "pc_csv": r"D:\Archivos de Programa\CapturasWS\PC1_TELETRAFICO\TEST 1\PC1T1_2.csv", "sv_csv": r"D:\Archivos de Programa\CapturasWS\SV\TEST 1\SVT1_P2.csv"},
            {"id": "P3", "nombre": "Prueba 3", "pc_csv": r"D:\Archivos de Programa\CapturasWS\PC1_TELETRAFICO\TEST 1\PC1T1_3.csv", "sv_csv": r"D:\Archivos de Programa\CapturasWS\SV\TEST 1\SVT1_P3.csv"}
        ]
    },
    {
        "equipo": "PC3",
        "ip": "192.168.0.186",  # <--- CAMBIAR AQUI LA IP DE PC3
        "archivo_salida": "TEST_1_PC3_Resultados.xlsx",
        "pruebas": [
            {"id": "P1", "nombre": "Prueba 1", "pc_csv": r"D:\Archivos de Programa\CapturasWS\PC3_TELETRAFICO\TEST 1\PC3T1_1.csv", "sv_csv": r"D:\Archivos de Programa\CapturasWS\SV\TEST 1\SVT1_P1.csv"},
            {"id": "P2", "nombre": "Prueba 2", "pc_csv": r"D:\Archivos de Programa\CapturasWS\PC3_TELETRAFICO\TEST 1\PC3T1_2.csv", "sv_csv": r"D:\Archivos de Programa\CapturasWS\SV\TEST 1\SVT1_P2.csv"},
            {"id": "P3", "nombre": "Prueba 3", "pc_csv": r"D:\Archivos de Programa\CapturasWS\PC3_TELETRAFICO\TEST 1\PC3T1_3.csv", "sv_csv": r"D:\Archivos de Programa\CapturasWS\SV\TEST 1\SVT1_P3.csv"}
        ]
    }
]

# ============================================================
# FUNCION DE PROCESAMIENTO CENTRAL
# ============================================================
def procesar_captura(pc_path, sv_path, ip_cliente):
    try:
        pc_df = pd.read_csv(pc_path)
        sv_df = pd.read_csv(sv_path)
    except FileNotFoundError as e:
        print(f"    [!] Archivo no encontrado: {e.filename}")
        return pd.DataFrame()

    for df in [pc_df, sv_df]:
        for col in ["tcp.seq", "tcp.ack", "tcp.len"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        df["frame.time"] = pd.to_datetime(df["frame.time"], errors="coerce")

    # Aislar unicamente el trafico entre el Servidor y el PC actual
    sv_aislado = sv_df[
        ((sv_df["ip.src"] == ip_cliente) & (sv_df["ip.dst"] == IP_SV)) |
        ((sv_df["ip.src"] == IP_SV) & (sv_df["ip.dst"] == ip_cliente))
    ].copy()

    datos_cliente = pc_df[
        (pc_df["ip.src"] == ip_cliente) &
        (pc_df["ip.dst"] == IP_SV) &
        (pc_df["tcp.len"] > 0)
    ].copy()

    grupos = datos_cliente.groupby(["tcp.seq", "tcp.len"], sort=False)
    resultados = []
    contador = 0

    for (seq, length), grupo in grupos:
        if contador >= MAX_SEGMENTOS:
            break

        grupo = grupo.sort_values("frame.time")
        A = grupo.iloc[0]["frame.time"] 

        seq = int(seq)
        length = int(length)
        ack_esperado = seq + length

        segmentos_sv = sv_aislado[
            (sv_aislado["ip.src"] == ip_cliente) &
            (sv_aislado["ip.dst"] == IP_SV) &
            (sv_aislado["tcp.seq"] == seq) &
            (sv_aislado["tcp.len"] == length) &
            (sv_aislado["frame.time"] >= A)
        ].copy()

        if len(segmentos_sv) == 0: continue
        B = segmentos_sv.sort_values("frame.time").iloc[0]["frame.time"]

        acks_sv = sv_aislado[
            (sv_aislado["ip.src"] == IP_SV) &
            (sv_aislado["ip.dst"] == ip_cliente) &
            (sv_aislado["tcp.ack"] == ack_esperado) &
            (sv_aislado["frame.time"] > B)
        ].copy()

        if len(acks_sv) == 0: continue
        ack_sv = acks_sv.sort_values("frame.time").iloc[0]
        C = ack_sv["frame.time"]
        ack_encontrado = int(ack_sv["tcp.ack"])
        ack_len = int(ack_sv["tcp.len"])

        paquetes_intermedios = sv_aislado[
            (sv_aislado["ip.src"] == ip_cliente) &
            (sv_aislado["ip.dst"] == IP_SV) &
            (sv_aislado["tcp.len"] > 0) &
            (sv_aislado["frame.time"] > B) &
            (sv_aislado["frame.time"] <= C)
        ]
        num_intermedios = len(paquetes_intermedios)

        acks_pc = pc_df[
            (pc_df["ip.src"] == IP_SV) &
            (pc_df["ip.dst"] == ip_cliente) &
            (pc_df["tcp.ack"] == ack_encontrado) &
            (pc_df["frame.time"] > A)
        ].copy()

        if len(acks_pc) == 0: continue
        D = acks_pc.sort_values("frame.time").iloc[0]["frame.time"]

        # Filtros estrictos de redes (aseguran que no haya desfases de reloj)
        RTT_s = (D - A).total_seconds()
        if RTT_s <= 0: continue

        respuesta_sv_s = (C - B).total_seconds()
        if respuesta_sv_s < 0: continue

        viaje_red_s = RTT_s - respuesta_sv_s
        if viaje_red_s < 0: continue

        c_s_s = viaje_red_s / 2
        offset_s = (B - A).total_seconds() - c_s_s
        s_c_real_s = (D - C).total_seconds() + offset_s
        total_descompuesto_s = c_s_s + respuesta_sv_s + s_c_real_s

        # Clasificacion
        if ack_len == 0 and respuesta_sv_s <= UMBRAL_INSTANTANEO_S:
            if num_intermedios == 0 and ack_encontrado == ack_esperado:
                tipo_ack = "ACK Individual"
            else:
                tipo_ack = "ACK Acumulativo"
        elif respuesta_sv_s > 0.020:
            tipo_ack = "Delayed ACK (Timer)"
        else:
            tipo_ack = "ACK con Datos (App/Ráfaga)"

        resultados.append({
            "SEQ": seq, "LEN": length, "ACK": ack_encontrado,
            "Tipo_ACK": tipo_ack,
            "Tiempo_C_S_s": c_s_s, "Respuesta_SV_s": respuesta_sv_s,
            "Tiempo_S_C_s": s_c_real_s, "Tiempo_Total_s": total_descompuesto_s
        })
        contador += 1

    return pd.DataFrame(resultados)


# ============================================================
# MOTOR DE EJECUCION
# ============================================================
print("=" * 80)
print("     INICIANDO PROCESAMIENTO MASIVO DE TOPOLOGIA (PC1 y PC3)")
print("=" * 80)

borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
columnas_decimales = ["Tiempo_C_S_s", "Respuesta_SV_s", "Tiempo_S_C_s", "Viaje_Red_Neto_s", "Tiempo_Total_RTT_s", "Tiempo_Total_s"]

for pc in TOPOLOGIA:
    nombre_equipo = pc["equipo"]
    ip_equipo = pc["ip"]
    archivo_out = pc["archivo_salida"]
    
    print(f"\n[{nombre_equipo}] Procesando capturas (IP: {ip_equipo})...")
    
    resultados_por_prueba = {}
    resumen_filas = []

    for prueba in pc["pruebas"]:
        p_id = prueba["id"]
        p_nombre = prueba["nombre"]
        print(f"  -> Analizando {p_nombre}...")
        
        df_prueba = procesar_captura(prueba["pc_csv"], prueba["sv_csv"], ip_equipo)
        resultados_por_prueba[p_id] = df_prueba
        
        if df_prueba.empty:
            continue

        df_inst = df_prueba[df_prueba["Tipo_ACK"] == "ACK Individual"]
        
        total_pkts = len(df_prueba)
        inst_pkts = len(df_inst)
        
        if inst_pkts > 0:
            c_s_mean = df_inst["Tiempo_C_S_s"].mean()
            r_sv_mean = df_inst["Respuesta_SV_s"].mean()
            s_c_mean = df_inst["Tiempo_S_C_s"].mean()
            red_mean = c_s_mean + s_c_mean
            rtt_mean = df_inst["Tiempo_Total_s"].mean()
        else:
            c_s_mean = r_sv_mean = s_c_mean = red_mean = rtt_mean = 0.0

        resumen_filas.append({
            "Prueba": p_nombre, "Total_Segmentos": total_pkts, "Muestras_Instantaneas": inst_pkts,
            "Porcentaje_Instantaneo": f"{(inst_pkts/total_pkts*100):.2f}%" if total_pkts > 0 else "0%",
            "Tiempo_C_S_s": c_s_mean, "Respuesta_SV_s": r_sv_mean,
            "Tiempo_S_C_s": s_c_mean, "Viaje_Red_Neto_s": red_mean, "Tiempo_Total_RTT_s": rtt_mean
        })

    if not resumen_filas:
        print(f"  [!] No se generaron datos validos para {nombre_equipo}. Saltando exportación.")
        continue

    # Guardar Excel para este PC
    df_resumen = pd.DataFrame(resumen_filas)
    with pd.ExcelWriter(archivo_out, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen_Comparativo", index=False)
        for p_id in ["P1", "P2", "P3"]:
            if p_id in resultados_por_prueba and not resultados_por_prueba[p_id].empty:
                df_p = resultados_por_prueba[p_id]
                df_inst = df_p[df_p["Tipo_ACK"] == "ACK Individual"].drop(columns=["Tipo_ACK"])
                df_otros = df_p[df_p["Tipo_ACK"] != "ACK Individual"]
                df_inst.to_excel(writer, sheet_name=f"Instantaneos_{p_id}", index=False)
                df_otros.to_excel(writer, sheet_name=f"Otros_{p_id}", index=False)

    # Formateo Visual
    wb = load_workbook(archivo_out)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2: continue
        
        color_header = "002060" if sheet_name == "Resumen_Comparativo" else ("1F4E78" if "Instantaneos" in sheet_name else "595959")
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor=color_header)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde

        indices = {cell.value: cell.column for cell in ws[1]}
        for col_name in columnas_decimales:
            if col_name in indices:
                col_idx = indices[col_name]
                for fila in range(2, ws.max_row + 1):
                    ws.cell(fila, col_idx).number_format = "0.000000"

        for columna in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in columna)
            letra = get_column_letter(columna[0].column)
            ws.column_dimensions[letra].width = max(max_len + 3, 16)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 25

    wb.save(archivo_out)
    print(f"  [✓] Guardado exitosamente: {archivo_out}")

print("\n" + "=" * 80)
print("             ANALISIS FINALIZADO COMPLETAMENTE")
print("=" * 80 + "\n")